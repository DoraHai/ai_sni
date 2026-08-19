"""真 xlsx 报告导出（区别于现有 xls=HTML table 方案）。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return round(v, 4)
    return v


def _write_table(ws, headers, rows, start_row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
    for col_idx, header in enumerate(headers, start=1):
        max_len = (
            max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows])
            if rows
            else len(str(header))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def build_report_workbook(report: dict) -> BytesIO:
    """
    report: get_analysis_report() / get_monthly_report() 的完整返回值，
            即 {"data": {...}, "narrative": {...} | None}
    返回内存 xlsx 二进制流。
    """
    data = report["data"]
    narrative = report.get("narrative") or {}
    tenant = data["tenant"]
    period = data["period"]

    wb = Workbook()

    ws = wb.active
    ws.title = "概览"
    ws.append([f"{tenant['name']} 投放分析报告"])
    ws["A1"].font = TITLE_FONT
    ws.append(
        [
            (
                f"周期：{period['start_date']} ~ {period['end_date']}"
                f"（共 {period['days']} 天，投放 {period['active_days']} 天）"
            )
        ]
    )
    ws.append([])

    if narrative.get("summary"):
        ws.append(["AI 总览"])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        ws.append([narrative["summary"]])
        ws.append([])

    kpi = data["kpi"]
    kpi_labels = [
        ("cost", "消费"),
        ("click", "点击"),
        ("impression", "展现"),
        ("cpc", "CPC"),
        ("ctr", "CTR"),
    ]
    _write_table(
        ws,
        ["指标", "本期", "上期", "环比(%)"],
        [
            [label, kpi[key]["current"], kpi[key]["previous"], kpi[key]["change_pct"]]
            for key, label in kpi_labels
        ],
        start_row=ws.max_row + 1,
    )

    budget = data["budget"]
    ws.append([])
    ws.append(["预算", "月度预算", budget.get("monthly_budget")])
    ws.append(["", "本月已消费", budget.get("month_cost")])
    ws.append(["", "本期消费", budget.get("period_cost")])
    ws.append(["", "预算使用比例(%)", budget.get("usage_pct")])

    alerts = data.get("alerts_review") or {}
    ws.append([])
    ws.append(["异常处置", "状态", "数量"])
    for status, count in alerts.items():
        ws.append(["", status, count])

    operations = data.get("operations") or {}
    ws.append([])
    ws.append(["优化操作", "总次数", operations.get("total")])
    for level, count in (operations.get("by_level") or {}).items():
        ws.append(["", f"层级-{level}", count])
    ws.append(["", "超20%调整上限次数", operations.get("over_limit")])
    ws.append(["", "AI建议采纳数", operations.get("ai_suggestions_adopted")])

    plans = narrative.get("next_period_plan") or narrative.get("next_month_plan") or []
    if plans:
        ws.append([])
        ws.append(["后续计划"])
        for idx, item in enumerate(plans, 1):
            ws.append([f"计划{idx}", item])

    ws2 = wb.create_sheet("日趋势")
    _write_table(
        ws2,
        ["日期", "消费", "点击", "展现"],
        [
            [t["date"], t["cost"], t["click"], t["impression"]]
            for t in (data.get("trend") or [])
        ],
    )

    ws3 = wb.create_sheet("关键词分类")
    _write_table(
        ws3,
        ["分类", "消费", "点击", "展现", "CPC", "CTR", "消费占比(%)"],
        [
            [
                c["category_label"],
                c["cost"],
                c["click"],
                c["impression"],
                c["cpc"],
                c["ctr"],
                c["cost_share_pct"],
            ]
            for c in (data.get("by_category") or [])
        ],
    )

    ws4 = wb.create_sheet("TOP关键词")
    _write_table(
        ws4,
        ["关键词", "消费", "点击", "展现", "CPC", "CTR", "平均排名"],
        [
            [
                k["keyword"],
                k["cost"],
                k["click"],
                k["impression"],
                k["cpc"],
                k["ctr"],
                k["avg_rank"],
            ]
            for k in (data.get("top_keywords") or [])
        ],
    )

    ws5 = wb.create_sheet("设备分布")
    _write_table(
        ws5,
        ["设备", "消费", "点击", "展现", "CPC", "CTR", "消费占比(%)"],
        [
            [
                d["device"],
                d["cost"],
                d["click"],
                d["impression"],
                d["cpc"],
                d["ctr"],
                d["cost_share_pct"],
            ]
            for d in (data.get("device_split") or [])
        ],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
